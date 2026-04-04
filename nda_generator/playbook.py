from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class PlaybookIssue:
    """Une ligne du playbook de négociation."""

    nom: str
    preferred_position: str
    fallback_position: str
    preferred_wording: str


def load_playbook(path: str | Path) -> list[PlaybookIssue]:
    """Charge les issues depuis la feuille active (lignes sous l'en-tête NOM DE L'ISSUE)."""
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    def _is_issue_header(row: tuple) -> bool:
        if not row or not row[0]:
            return False
        a = str(row[0]).strip().upper().replace("\u2019", "'")
        b = str(row[1] or "").strip().upper()
        return "NOM" in a and "ISSUE" in a and "PREFERRED" in b

    header_idx = None
    for i, row in enumerate(rows):
        if _is_issue_header(row):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            f"Aucune ligne d'en-tête trouvée (colonne « NOM DE L'ISSUE ») dans {path}"
        )

    issues: list[PlaybookIssue] = []
    for row in rows[header_idx + 1 :]:
        if not row or row[0] is None:
            continue
        nom = str(row[0]).strip()
        if not nom:
            continue
        issues.append(
            PlaybookIssue(
                nom=nom,
                preferred_position=str(row[1] or "").strip(),
                fallback_position=str(row[2] or "").strip(),
                preferred_wording=str(row[3] or "").strip(),
            )
        )
    return issues
